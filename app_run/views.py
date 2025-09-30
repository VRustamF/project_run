from openpyxl.styles.builtins import total
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination

from django.db.models import Count, Q, Sum, Min, Max, Avg
from django.conf import settings
from django.shortcuts import get_object_or_404

from django_filters.rest_framework import DjangoFilterBackend

from .serializers import (
    RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, PositionSerializer,
    CollectibleItemSerializer, SubscribeSerializer, AthletesSubscriptionsSerializer, CoachFollowersSerializer,
    AnalyticsSerializer
)
from .models import Run, User, AthleteInfo, Challenge, Position, CollectibleItem, Subscribe

from haversine import haversine, Unit
from openpyxl import load_workbook



@api_view(['GET'])
def company_details(request):
    details = {'company_name': settings.COMPANY_NAME,
               'slogan': settings.SLOGAN,
               'contacts': settings.CONTACTS}
    return Response(details)



class BasePagination(PageNumberPagination):
    page_size_query_param = 'size'



class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.all().select_related('athlete')
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    pagination_class = BasePagination



class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.annotate(runs_finished=Count('runs', filter=Q(runs__status='finished')), rating=Avg('coach__rating'))
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']
    pagination_class = BasePagination

    def get_queryset(self):
        qs = self.queryset
        type = self.request.query_params.get('type', None)
        if type == 'coach':
            qs = qs.filter(is_staff=True, is_superuser=False)
        elif type == 'athlete':
            qs = qs.filter(is_staff=False, is_superuser=False)
        else:
            qs = qs.filter(is_superuser=False)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            user = self.get_object()

            if user.is_staff:
                return CoachFollowersSerializer
            else:
                return AthletesSubscriptionsSerializer

        return super().get_serializer_class()



class StartAPIView(APIView):
    serializer_class = RunSerializer

    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == Run.Status.INIT:
            run.status = Run.Status.IN_PROGRESS
            run.save()
            serializer = self.serializer_class(run)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)



class StopAPIView(APIView):
    serializer_class = RunSerializer

    def check_challenge_ten_runs(self, athlete):
        finished = athlete.runs.filter(status=Run.Status.FINISHED).count()
        if finished == 10:
            self.assign_challenge(athlete, full_name='Сделай 10 Забегов!')

    def check_challenge_fifty_km(self, athlete):
        total_km = athlete.runs.filter(status=Run.Status.FINISHED).aggregate(total=Sum('distance'))['total']
        if total_km and total_km >= 50:
            self.assign_challenge(athlete, full_name='Пробеги 50 километров!')

    def chech_chellenge_twokm_tenmin(self, run):
        if run.distance >= 2 and run.run_time_seconds <= 600:
            athlete = run.athlete
            self.assign_challenge(athlete, full_name='2 километра за 10 минут!')

    def assign_challenge(self, athlete, full_name):
        Challenge.objects.create(full_name=full_name, athlete=athlete)

    def distance_calculation(self, run):
        queryset = run.position.all()
        stack = []
        distance = 0
        for positions in queryset:
            current_position = (positions.latitude, positions.longitude)
            if stack:
                distance += haversine(stack[-1], current_position)
            stack.append(current_position)
        return distance

    def run_time(self, run):
        agg_positions = run.position.all().aggregate(
            first_created_pos=Min('date_time'),
            last_created_pos=Max('date_time')
        )
        first_pos, last_pos = agg_positions['first_created_pos'], agg_positions['last_created_pos']
        if first_pos and last_pos:
            run.run_time_seconds = int((last_pos - first_pos).total_seconds())
            run.save()

    def post(self, request, run_id):
        run = get_object_or_404(Run.objects.select_related('athlete').prefetch_related('position'), id=run_id)
        if run.status == Run.Status.IN_PROGRESS:
            run.status = Run.Status.FINISHED
            run.distance = self.distance_calculation(run)
            avg_speed = run.position.all().aggregate(avg_speed=Avg('speed'))['avg_speed']
            run.speed = round(avg_speed if avg_speed else 0, 2)
            run.save()
            self.check_challenge_ten_runs(run.athlete)
            self.check_challenge_fifty_km(run.athlete)
            self.run_time(run)
            self.chech_chellenge_twokm_tenmin(run)
            serializer = self.serializer_class(run)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class AthleteInfoAPIView(APIView):
    serializer_class = AthleteInfoSerializer

    def get(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        athlete, created = AthleteInfo.objects.get_or_create(user=user, defaults={'goals': '', 'weight': 0})
        serializer = self.serializer_class(athlete)

        return Response(data=serializer.data, status=status.HTTP_200_OK)

    def put(self, request, user_id):
        user = get_object_or_404(User, id=user_id)

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        athlete, created = AthleteInfo.objects.update_or_create(user=user, defaults=serializer.validated_data)
        response_serializer = self.serializer_class(athlete)
        status_code = status.HTTP_201_CREATED if created else status.HTTP_200_OK

        return Response(data=response_serializer.data, status=status_code)


class ChallengesViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Challenge.objects.all().select_related('athlete')
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        qs = self.queryset
        athlete_id = self.request.query_params.get('athlete', None)
        if athlete_id:
            qs = qs.filter(athlete=athlete_id)
        return qs



class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all().select_related('run')
    serializer_class = PositionSerializer

    def check_item(self, item, position_instance):
        current_position = (position_instance.latitude, position_instance.longitude)
        item_position = (item.latitude, item.longitude)
        distance = haversine(item_position, current_position, unit=Unit.METERS)
        return distance < 100


    def get_queryset(self):
        qs = self.queryset
        run_id = self.request.query_params.get('run', None)
        if run_id:
            return qs.filter(run=run_id)
        return qs

    def perform_create(self, serializer):
        qs = self.queryset.filter(run__id=serializer.validated_data['run'].id)
        last_pos = qs.last()

        speed, final_distance = 0, 0

        if last_pos:
            last_cords = (last_pos.latitude, last_pos.longitude)
            last_pos_time = last_pos.date_time

            current_cords = (serializer.validated_data['latitude'], serializer.validated_data['longitude'])
            current_pos_time = (serializer.validated_data['date_time'])

            distance = haversine(last_cords, current_cords, unit=Unit.METERS)
            time = int((current_pos_time - last_pos_time).total_seconds())
            speed = round(distance / time, 2)
            final_distance = (last_pos.distance or 0) + round(haversine(last_cords, current_cords), 2)

        position_instance = serializer.save(speed=speed, distance=final_distance)
        athlete = position_instance.run.athlete
        items = CollectibleItem.objects.all()
        for item in items:
            if self.check_item(item, position_instance):
                item.athletes.add(athlete)



class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer



class CollectibleItemAPIView(APIView):
    serializer_class = CollectibleItemSerializer

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(file)
        sheet = workbook.active
        invalid = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = {
                'name': row[0],
                'uid': row[1],
                'value': row[2],
                'latitude': row[3],
                'longitude': row[4],
                'picture': row[5],
            }
            serializer = self.serializer_class(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                invalid.append(list(row))
        return Response(data=invalid)



class SubscribeAPIView(APIView):
    serializer_class = SubscribeSerializer

    def post(self, request, coach_id=None):
        queryset = User.objects.filter(is_superuser=False)
        coach = get_object_or_404(queryset, id=coach_id)
        if not coach.is_staff:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        athlete_id = request.data.get('athlete')
        try:
            athlete = queryset.get(id=athlete_id)
        except User.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        if athlete.is_staff:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        subscribe, created = Subscribe.objects.get_or_create(coach=coach, athlete=athlete)

        if created:
            serializer = self.serializer_class(subscribe)
            return Response(data=serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)



class ChallengesSummaryViewSet(viewsets.ReadOnlyModelViewSet):
    def list(self, request, *args, **kwargs):
        queryset = Challenge.objects.all().select_related('athlete')

        result = {}
        for challenge in queryset:
            athlete = challenge.athlete
            challenge_name = challenge.full_name
            athlete_info = {
                'id': athlete.id,
                'full_name': f'{athlete.first_name} {athlete.last_name}',
                'username': athlete.username
            }
            result.setdefault(challenge_name, {'name_to_display': challenge_name, 'athletes':[]})
            result[challenge_name]['athletes'].append(athlete_info)

        return Response(list(result.values()))



class RateCoachAPIView(APIView):

    def post(self, request, coach_id=None):
        subscribes = Subscribe.objects.all().select_related('athlete').select_related('coach')
        athlete_id = request.data.get('athlete')
        rating = request.data.get('rating')
        if not athlete_id:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        users = User.objects.filter(Q(id=coach_id, is_staff=True) | Q(id=athlete_id, is_staff=False))

        coach = get_object_or_404(users, id=coach_id, is_staff=True)

        try:
            athlete = users.get(id=athlete_id, is_staff=False)
        except User.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        subscribe = subscribes.filter(athlete=athlete, coach=coach).first()

        if subscribe and isinstance(rating, int) and 0 < rating <= 5:
            subscribe.rating = rating
            subscribe.save()
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class AnalyticsAPIView(APIView):

    def get(self, request, coach_id):
        athletes = (
            User.objects.filter(athlete__coach_id=coach_id)
            .annotate(
                longest_run=Max('runs__distance'),
                total_distance=Sum('runs__distance'),
                avg_speed=Avg('runs__speed')
            )
        )

        if not athletes.exists():
            return Response(data={}, status=status.HTTP_200_OK)

        longest_run_value = athletes.aggregate(Max('longest_run'))['longest_run__max']
        total_run_value = athletes.aggregate(Max('total_distance'))['total_distance__max']
        speed_avg_value = athletes.aggregate(Max('avg_speed'))['avg_speed__max']

        longest = athletes.filter(longest_run=longest_run_value).first()
        total = athletes.filter(total_distance=total_run_value).first()
        fastest = athletes.filter(avg_speed=speed_avg_value).first()

        analytics_data = {
            'longest_run_user': longest.id if longest else None,
            'longest_run_value': longest_run_value,

            'total_run_user': total.id if total else None,
            'total_run_value': total_run_value,

            'speed_avg_user': fastest.id if fastest else None,
            'speed_avg_value': speed_avg_value,
        }

        return Response(data=analytics_data, status=status.HTTP_200_OK)